"""Parses RankingsTiersMarketScore_2026.xlsx into per-player records.

The workbook has two sheets:
  - "Rankings and Tiers": one row per player (Overall, Player, Position,
    Pos Rank, Tier, Auction value).
  - "Market Score": four side-by-side position blocks (QB/RB/WR/TE), each
    a 4-column (Rank, Player, ADP, Market Score) table with a blank spacer
    column between blocks.

Records from both sheets are merged by (player name, position) into one
dict per player. Player names are used as-is from the source; reconciling
them against ESPN/Yahoo IDs is the job of the separate player-ID crosswalk.
"""

import openpyxl

MARKET_SCORE_BLOCKS = [
    ("QB", 0),
    ("RB", 5),
    ("WR", 10),
    ("TE", 15),
]


def parse_rankings_and_tiers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Rankings and Tiers"]
    records = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        overall, player, position, pos_rank, tier, auction_value = row[:6]
        if player is None:
            continue
        records[(player, position)] = {
            "player": player,
            "position": position,
            "overall_rank": overall,
            "pos_rank": pos_rank,
            "tier": tier,
            "auction_value": auction_value,
            "adp": None,
            "market_score": None,
        }
    return records


def parse_market_score(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Market Score"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    updates = {}
    for position, col_offset in MARKET_SCORE_BLOCKS:
        for row in rows:
            block = row[col_offset:col_offset + 4]
            if len(block) < 4:
                continue
            rank, player, adp, market_score = block
            if player is None:
                continue
            updates[(player, position)] = {
                "market_score_pos_rank": rank,
                "adp": adp,
                "market_score": market_score,
            }
    return updates


def load_players(path):
    records = parse_rankings_and_tiers(path)
    market_updates = parse_market_score(path)

    unmatched = []
    for key, update in market_updates.items():
        if key in records:
            records[key].update(update)
        else:
            unmatched.append(key)

    return list(records.values()), unmatched


if __name__ == "__main__":
    import sys
    import json

    path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\T991158\Downloads\RankingsTiersMarketScore_2026.xlsx"
    players, unmatched = load_players(path)
    print(f"parsed {len(players)} players")
    if unmatched:
        print(f"WARNING: {len(unmatched)} Market Score entries had no Rankings/Tiers match:")
        for key in unmatched:
            print("  ", key)
    print(json.dumps(players[:5], indent=2))
